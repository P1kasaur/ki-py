import pandas as pd
import requests
import urllib3.request #pip install urllib3 
import time
from bs4 import BeautifulSoup
import re
import openpyxl
import os


filein = input('File input: ').lstrip('"').rstrip('"')
userin = input('What IP Address?\n1:Source\n2:Destination\nChoose: ')

ipnum = {
	'1' : 'Source IP Address',
	'2' : 'Destination IP Address'
}

ipcoun = {
	'1' : 'Source Country',
	'2' : 'Destination Country'
}

# making data frame from csv file 
data = pd.read_csv(filein)
 
# sorting by first name 
data.sort_values(ipnum[userin], inplace = True)
data.drop_duplicates(subset =ipnum[userin], keep = 'last' , inplace = True) 

a = data.loc[:, [ipnum[userin],ipcoun[userin]]] 


writer = pd.ExcelWriter(r'C:\Users\luigic\Desktop\pyxl\result.xlsx', engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object.
a.to_excel(writer, sheet_name='Sheet1')

# Close the Pandas Excel writer and output the Excel file.
writer.save()
#print a


filepath = r'C:\Users\luigic\Desktop\pyxl\result.xlsx'
wb = openpyxl.load_workbook(filepath)
sheet = wb.active

for i in range(2 , sheet.max_row + 1):
	ip = str(sheet.cell(row=i,column=2).value)

	url = 'https://iplogger.org/ip-lookup/?d=' + ip
	response = requests.get(url)

	soup = str(BeautifulSoup(response.text, 'html.parser'))

	isp = re.findall('ISP"\/> .*',soup)

	isp = isp[0].replace('ISP"/> ','')
	isp = re.sub('\s\[\w.+?]','',isp)
	sheet.cell(row=i,column=4).value = isp
	
	country = re.findall('Country"\/> .*',soup)
	country = country[0].replace('Country"/> ','')
	sheet.cell(row=i,column=5).value = country	
	time.sleep(1)
	print('IP ' + str(i - 1) + ' done')
wb.save(filepath)
print('done whois query')


'''start = time.time()
a = range(100000)
b = [i*2 for i in a]
end = time.time()
print(end - start)'''

